import re
import json
import unicodedata
import pandas as pd
from pypdf import PdfReader
from openpyxl.styles import Alignment, Font


TXT_PATH = "orbi.txt"
PDF_PATH = "cevaporbital.pdf"
OUTPUT_XLSX = "sinav_sonuclari_orbital_tyt1.xlsx"

TESTS = [
    ("Türkçe", 40),
    ("Sosyal", 20),
    ("Matematik", 40),
    ("Fen", 20),
]
DEFAULT_TEST_LENGTHS = dict(TESTS)

# TYT puanı için kurum çıktısına kalibre edilmiş katsayılar.
# (11.SINIF SÜREÇ ANALİZİ 1 gerçek sonuç tablosundan türetilmiştir.)
TYT_BASE_SCORE = 144.9496405
TYT_COEFFICIENTS = {
    "Türkçe": 2.88790178,
    "Sosyal": 2.91602531,
    "Matematik": 2.90504068,
    "Fen": 3.12600659,
}

SOZ_BASE_SCORE = 111.356738
SOZ_COEFFICIENTS = {
    "Türkçe": 6.48254985,
    "Sosyal": 6.62280338,
    "Matematik": -0.183190932,
    "Fen": -0.00771334513,
}

SAY_BASE_SCORE = 109.61517569
SAY_COEFFICIENTS = {
    "Türkçe": -0.41844955,
    "Sosyal": 0.31249024,
    "Matematik": 6.6088891,
    "Fen": 6.29762888,
}

EA_BASE_SCORE = 114.711013
EA_COEFFICIENTS = {
    "Türkçe": 6.82709757,
    "Sosyal": -0.00002789615,
    "Matematik": 6.0159298,
    "Fen": 0.00003545567,
}

# AYT-benzeri oturumlar için yaklaşık kurum puanı.
AYT_BASE_SCORE = 100.0
AYT_COEFFICIENTS = {
    "Türkçe": 2.0,
    "Sosyal": 2.0,
    "Matematik": 2.0,
    "Fen": 2.0,
}


def normalize_answers(s):
    return "".join(ch if ch in "ABCDE" else " " for ch in s)


def to_ascii_upper(s):
    n = unicodedata.normalize("NFKD", s)
    n = "".join(ch for ch in n if not unicodedata.combining(ch))
    return n.upper()


def norm_test_name(line):
    t = to_ascii_upper(line)
    if "TURK DILI VE EDEBIYATI" in t or "TURK DILI EDEBIYATI" in t:
        return "Türkçe"
    if "TURKCE" in t:
        return "Türkçe"
    if "SOSYAL BILIMLER" in t:
        return "Sosyal"
    if "SOSYAL" in t:
        return "Sosyal"
    if "TEMEL MATEMATIK" in t or "MATEMATIK" in t:
        return "Matematik"
    if "FEN BILIMLERI" in t:
        return "Fen"
    if "FEN" in t:
        return "Fen"
    return None


def _subjects_for_tyt_table_row(q_num):
    if q_num <= 20:
        return ["Türkçe", "Sosyal", "Matematik", "Fen"]
    if q_num <= 25:
        return ["Türkçe", "Sosyal", "Matematik"]
    return ["Türkçe", "Matematik"]


def _parse_side_by_side_table_line(line):
    tokens = line.split()
    if not tokens or not tokens[0].isdigit():
        return None

    q_num = int(tokens[0])
    if q_num < 1 or q_num > 40:
        return None

    split_idx = None
    for idx in range(1, len(tokens)):
        if tokens[idx] == str(q_num):
            split_idx = idx
            break
    if split_idx is None:
        return None

    subjects = _subjects_for_tyt_table_row(q_num)
    parsed = {"A": {}, "B": {}}
    for booklet, answers in (
        ("A", tokens[1:split_idx]),
        ("B", tokens[split_idx + 1 :]),
    ):
        for subject, ans in zip(subjects, answers):
            ans = to_ascii_upper(ans)
            if ans in "ABCDE":
                parsed[booklet][(subject, q_num)] = ans
    return parsed


def _detect_side_by_side_table(page_text):
    up = to_ascii_upper(page_text)
    if "SORU NO" not in up or "TURKCE" not in up or "MATEMATIK" not in up:
        return False
    if not ("A KITAP" in up and "B KITAP" in up):
        return False

    data_rows = 0
    for line in page_text.splitlines():
        parsed = _parse_side_by_side_table_line(line.strip())
        if parsed and parsed["A"] and parsed["B"]:
            data_rows += 1
    return data_rows >= 10


def _parse_side_by_side_table_pdf(reader):
    keys = {"A": {name: {} for name, _ in TESTS}, "B": {name: {} for name, _ in TESTS}}
    found_rows = 0

    for page in reader.pages:
        page_text = page.extract_text() or ""
        if not _detect_side_by_side_table(page_text):
            continue

        for line in page_text.splitlines():
            parsed = _parse_side_by_side_table_line(line.strip())
            if not parsed:
                continue
            found_rows += 1
            for booklet in ["A", "B"]:
                for (subject, q_num), ans in parsed[booklet].items():
                    keys[booklet][subject][q_num] = ans

    if found_rows < 10:
        return None

    final = {"A": {}, "B": {}}
    for booklet in ["A", "B"]:
        for test_name, default_len in TESTS:
            section = keys[booklet][test_name]
            max_q = max(section.keys()) if section else 0
            length = max_q if max_q > 0 else default_len
            final[booklet][test_name] = "".join(section.get(i, " ") for i in range(1, length + 1))
    return final


def parse_pdf_keys(pdf_path):
    reader = PdfReader(pdf_path)
    side_by_side_keys = _parse_side_by_side_table_pdf(reader)
    if side_by_side_keys and key_quality(side_by_side_keys) >= 160:
        return side_by_side_keys

    text = ""
    for page in reader.pages:
        text += (page.extract_text() or "") + "\n"

    keys = {"A": {}, "B": {}}
    booklet = None
    current_test = None
    # Sequence-style PDF'lerde soru numarası olmadan harf akışı geliyor.
    # Bu durumda üst sınırlar TYT/AYT ortak kapsayacak şekilde geniş tutulur.
    test_counts = {
        "Türkçe": 40,
        "Sosyal": 46,      # AYT Sosyal-2: 46
        "Matematik": 40,
        "Fen": 40,         # AYT Fen: 40
    }

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    first_test_idx = next((i for i, ln in enumerate(lines) if norm_test_name(ln)), -1)
    a_header_idx = next((i for i, ln in enumerate(lines) if "A KITAP" in to_ascii_upper(ln) or re.search(r"\(A\)", to_ascii_upper(ln))), -1)
    b_header_idx = next((i for i, ln in enumerate(lines) if "B KITAP" in to_ascii_upper(ln) or re.search(r"\(B\)", to_ascii_upper(ln))), -1)
    alternating_mode = (
        first_test_idx != -1
        and a_header_idx != -1
        and b_header_idx != -1
        and a_header_idx < first_test_idx
        and b_header_idx < first_test_idx
    )
    section_occurrence = {name: 0 for name, _ in TESTS}

    def add_by_sequence(booklet_name, test_name, letters):
        d = keys[booklet_name].setdefault(test_name, {})
        limit = test_counts[test_name]
        next_q = len(d) + 1
        for ch in letters:
            if next_q > limit:
                break
            d[next_q] = ch
            next_q += 1

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue

        up = to_ascii_upper(line)
        if not alternating_mode and ("A KITAP" in up or re.search(r"\(A\)", up)):
            booklet = "A"
            current_test = None
            continue
        if not alternating_mode and ("B KITAP" in up or re.search(r"\(B\)", up)):
            booklet = "B"
            current_test = None
            continue
        if alternating_mode and ("A KITAP" in up or "B KITAP" in up or re.search(r"\([AB]\)", up)):
            continue
        if not alternating_mode and booklet is None:
            continue

        test_name = norm_test_name(line)
        if test_name:
            current_test = test_name
            if alternating_mode:
                idx = section_occurrence[test_name]
                booklet = "A" if idx % 2 == 0 else "B"
                section_occurrence[test_name] += 1
            continue

        if current_test is None:
            continue

        # Format 1: "1. A 2. B ..."
        for q, ans in re.findall(r"(\d+)\.\s*([ABCDE])", line):
            keys[booklet].setdefault(current_test, {})[int(q)] = ans

        # Format 2 fallback: separate answer lines like "A D E C ..."
        if not re.search(r"\d", line):
            letters = re.findall(r"[ABCDE]", up)
            if letters:
                add_by_sequence(booklet, current_test, letters)

    # Format 3 fallback: chunk by booklet/test headings and stream letters.
    # This catches PDFs where numbering is broken or punctuation lost.
    if key_quality(keys) < 160:
        keys2 = {"A": {}, "B": {}}
        booklet2 = None
        test2 = None
        for line in lines:
            up = to_ascii_upper(line)
            if "A KITAP" in up or re.search(r"\(A\)", up):
                booklet2 = "A"
                test2 = None
                continue
            if "B KITAP" in up or re.search(r"\(B\)", up):
                booklet2 = "B"
                test2 = None
                continue
            if booklet2 is None:
                continue

            ntest = norm_test_name(line)
            if ntest:
                test2 = ntest
                continue
            if test2 is None:
                continue

            letters = re.findall(r"[ABCDE]", up)
            if letters:
                d = keys2[booklet2].setdefault(test2, {})
                limit = test_counts[test2]
                next_q = len(d) + 1
                for ch in letters:
                    if next_q > limit:
                        break
                    d[next_q] = ch
                    next_q += 1

        # Pick better parse by filled-answer quality.
        def to_final(kdict):
            out = {"A": {}, "B": {}}
            for b in ["A", "B"]:
                for test_name, default_len in TESTS:
                    d = kdict[b].get(test_name, {})
                    max_q = max(d.keys()) if d else 0
                    length = max_q if max_q > 0 else default_len
                    out[b][test_name] = "".join(d.get(i, " ") for i in range(1, length + 1))
            return out

        final1 = {"A": {}, "B": {}}
        for b in ["A", "B"]:
            for test_name, default_len in TESTS:
                d = keys[b].get(test_name, {})
                max_q = max(d.keys()) if d else 0
                length = max_q if max_q > 0 else default_len
                final1[b][test_name] = "".join(d.get(i, " ") for i in range(1, length + 1))
        final2 = to_final(keys2)

        if key_quality(final2) > key_quality(final1):
            return final2

    final_keys = {"A": {}, "B": {}}
    for booklet in ["A", "B"]:
        for test_name, default_len in TESTS:
            d = keys[booklet].get(test_name, {})
            max_q = max(d.keys()) if d else 0
            length = max_q if max_q > 0 else default_len
            key = "".join(d.get(i, " ") for i in range(1, length + 1))
            final_keys[booklet][test_name] = key

    return final_keys


def normalize_keys(keys):
    out = {"A": {}, "B": {}}
    for booklet in ["A", "B"]:
        for test_name, default_len in TESTS:
            raw = keys.get(booklet, {}).get(test_name, "")
            if isinstance(raw, dict):
                max_q = max(raw.keys()) if raw else 0
                length = max_q if max_q > 0 else default_len
                s = "".join(raw.get(i, " ") for i in range(1, length + 1))
            else:
                s = str(raw)
                length = len(s) if s.strip() else default_len
            s = normalize_answers(s)
            s = (s + (" " * length))[:length]
            out[booklet][test_name] = s
    return out


def save_standard_key_json(keys, path):
    tests = []
    for test_name, default_len in TESTS:
        length = (
            max(
                len(keys.get("A", {}).get(test_name, "")),
                len(keys.get("B", {}).get(test_name, "")),
            )
            or default_len
        )
        tests.append({"name": test_name, "length": length})

    payload = {
        "version": 1,
        "exam_type": "TYT",
        "tests": tests,
        "booklets": keys,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_standard_key_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return normalize_keys(data.get("booklets", {}))


def key_quality(keys):
    # Count non-space answers for TYT tests.
    total = 0
    for booklet in ["A", "B"]:
        for test_name, _ in TESTS:
            v = keys[booklet].get(test_name, "")
            if isinstance(v, dict):
                total += sum(1 for ch in v.values() if isinstance(ch, str) and ch in "ABCDE")
            else:
                total += sum(1 for ch in v if ch in "ABCDE")
    return total


def calculate_score(student_ans, key_str):
    dogru = 0
    yanlis = 0
    bos = 0

    for s, k in zip(student_ans, key_str):
        if k == " ":
            continue
        if s == " ":
            bos += 1
        elif s == k:
            dogru += 1
        else:
            yanlis += 1

    net = dogru - (yanlis / 4.0)
    return dogru, yanlis, bos, net


def get_test_lengths(keys):
    def observed_len(s):
        return len(s) if any(ch in "ABCDE" for ch in s) else 0

    return {
        test_name: (
            max(
                observed_len(keys.get("A", {}).get(test_name, "")),
                observed_len(keys.get("B", {}).get(test_name, "")),
            )
            or DEFAULT_TEST_LENGTHS[test_name]
        )
        for test_name, _ in TESTS
    }


def detect_exam_profile(lengths):
    signature = (
        lengths["Türkçe"],
        lengths["Sosyal"],
        lengths["Matematik"],
        lengths["Fen"],
    )
    if signature == (30, 30, 30, 30):
        return "school_30x4"
    if signature == (40, 20, 40, 20):
        return "tyt"
    if signature[0] == 40 and signature[2] == 40 and signature[3] == 20 and 20 <= signature[1] <= 25:
        return "tyt"
    return "ayt_like"


def answers_from_layout(line, s1, s2, o1, o2, lengths):
    tur_l = lengths["Türkçe"]
    sos_l = lengths["Sosyal"]
    mat_l = lengths["Matematik"]
    fen_l = lengths["Fen"]
    g1 = tur_l + sos_l
    g2 = mat_l + fen_l

    b1 = normalize_answers(line[s1:s1 + g1])
    b2 = normalize_answers(line[s2:s2 + g2])

    if o1 == 0:
        tur = b1[:tur_l]
        sos = b1[tur_l:tur_l + sos_l]
    else:
        sos = b1[:sos_l]
        tur = b1[sos_l:sos_l + tur_l]

    if o2 == 0:
        mat = b2[:mat_l]
        fen = b2[mat_l:mat_l + fen_l]
    else:
        fen = b2[:fen_l]
        mat = b2[fen_l:fen_l + mat_l]

    return {
        "Türkçe": tur,
        "Sosyal": sos,
        "Matematik": mat,
        "Fen": fen,
    }


def answers_from_four_blocks(line, starts, lengths):
    t_start, s_start, m_start, f_start = starts
    tur_l = lengths["Türkçe"]
    sos_l = lengths["Sosyal"]
    mat_l = lengths["Matematik"]
    fen_l = lengths["Fen"]
    return {
        "Türkçe": normalize_answers(line[t_start:t_start + tur_l]),
        "Sosyal": normalize_answers(line[s_start:s_start + sos_l]),
        "Matematik": normalize_answers(line[m_start:m_start + mat_l]),
        "Fen": normalize_answers(line[f_start:f_start + fen_l]),
    }


def extract_answers(line, layout, lengths):
    if layout["mode"] == "four_blocks":
        return answers_from_four_blocks(line, layout["starts"], lengths)
    return answers_from_layout(
        line,
        layout["s1"],
        layout["s2"],
        layout["o1"],
        layout["o2"],
        lengths,
    )


def _evaluate_layout(lines, keys, lengths, layout):
    totals = []
    valid = True
    for line in lines:
        if layout["mode"] == "two_blocks":
            g1 = lengths["Türkçe"] + lengths["Sosyal"]
            g2 = lengths["Matematik"] + lengths["Fen"]
            if len(line) < max(layout["s1"] + g1, layout["s2"] + g2):
                valid = False
                break
        else:
            t_start, s_start, m_start, f_start = layout["starts"]
            if len(line) < max(
                t_start + lengths["Türkçe"],
                s_start + lengths["Sosyal"],
                m_start + lengths["Matematik"],
                f_start + lengths["Fen"],
            ):
                valid = False
                break

        ans = extract_answers(line, layout, lengths)
        a = score_with_booklet(ans, keys["A"])
        b = score_with_booklet(ans, keys["B"])
        best = a if a["Toplam Net"] >= b["Toplam Net"] else b
        totals.append(best["Toplam Net"])

    if not valid or not totals:
        return None
    return (sum(totals) / len(totals), min(totals))


def _detect_layout_two_blocks(lines, keys, lengths):
    lengths = get_test_lengths(keys)
    g1 = lengths["Türkçe"] + lengths["Sosyal"]
    g2 = lengths["Matematik"] + lengths["Fen"]
    min_line_len = max(len(x) for x in lines) if lines else 0
    max_s1 = max(45, min(120, min_line_len - g1))
    min_s2 = 80
    max_s2 = max(min_s2, min(220, min_line_len - g2))

    # Auto detect answer blocks.
    candidates = []
    for s1 in range(45, max_s1 + 1):
        for s2 in range(min_s2, max_s2 + 1):
            for o1 in [0, 1]:
                for o2 in [0, 1]:
                    layout = {"mode": "two_blocks", "s1": s1, "s2": s2, "o1": o1, "o2": o2}
                    metrics = _evaluate_layout(lines, keys, lengths, layout)
                    if metrics is None:
                        continue
                    mean_total, min_total = metrics
                    candidates.append((mean_total, min_total, layout))

    if not candidates:
        return None

    # Prefer higher mean total; tie-breaker higher minimum.
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0]


def _detect_layout_four_blocks(lines, keys, lengths):
    if not lines:
        return None

    min_start = 45
    max_line_len = max(len(x) for x in lines)
    if max_line_len <= min_start:
        return None

    letters_per_col = [0] * max_line_len
    for line in lines:
        for i, ch in enumerate(line):
            if ch in "ABCDE":
                letters_per_col[i] += 1

    test_order = ["Türkçe", "Sosyal", "Matematik", "Fen"]
    test_lens = [lengths[t] for t in test_order]
    max_starts = [max_line_len - l for l in test_lens]
    if any(ms < min_start for ms in max_starts):
        return None

    window_scores = []
    for l in test_lens:
        prefix = [0]
        for v in letters_per_col:
            prefix.append(prefix[-1] + v)
        scores = [0] * max_line_len
        for s in range(min_start, max_line_len - l + 1):
            scores[s] = prefix[s + l] - prefix[s]
        window_scores.append(scores)

    gap = 2
    l1, l2, l3, l4 = test_lens
    s1_range = range(min_start, max_line_len - (l1 + l2 + l3 + l4 + 3 * gap) + 1)

    best4_from = [(-1, -1)] * (max_line_len + 2)
    best_score = -1
    best_start = -1
    for s in range(max_line_len - l4, min_start - 1, -1):
        score = window_scores[3][s]
        if score >= best_score:
            best_score = score
            best_start = s
        best4_from[s] = (best_score, best_start)

    best3_from = [(-1, -1, -1)] * (max_line_len + 2)
    best_score = -1
    best_s3 = -1
    best_s4 = -1
    for s3 in range(max_line_len - l3, min_start - 1, -1):
        s4_min = s3 + l3 + gap
        if s4_min > max_line_len - l4:
            best3_from[s3] = (best_score, best_s3, best_s4)
            continue
        s4_score, s4 = best4_from[s4_min]
        cand = window_scores[2][s3] + s4_score
        if cand >= best_score:
            best_score = cand
            best_s3 = s3
            best_s4 = s4
        best3_from[s3] = (best_score, best_s3, best_s4)

    best2_from = [(-1, -1, -1, -1)] * (max_line_len + 2)
    best_score = -1
    best_s2 = -1
    best_s3 = -1
    best_s4 = -1
    for s2 in range(max_line_len - l2, min_start - 1, -1):
        s3_min = s2 + l2 + gap
        if s3_min > max_line_len - l3:
            best2_from[s2] = (best_score, best_s2, best_s3, best_s4)
            continue
        s3_score, s3, s4 = best3_from[s3_min]
        cand = window_scores[1][s2] + s3_score
        if cand >= best_score:
            best_score = cand
            best_s2 = s2
            best_s3 = s3
            best_s4 = s4
        best2_from[s2] = (best_score, best_s2, best_s3, best_s4)

    best_candidate = None
    for s1 in s1_range:
        s2_min = s1 + l1 + gap
        if s2_min > max_line_len - l2:
            continue
        s2_score, s2, s3, s4 = best2_from[s2_min]
        if s2 == -1:
            continue
        starts = (s1, s2, s3, s4)
        layout = {"mode": "four_blocks", "starts": starts}
        metrics = _evaluate_layout(lines, keys, lengths, layout)
        if metrics is None:
            continue
        mean_total, min_total = metrics
        cand = (mean_total, min_total, layout)
        if best_candidate is None or (cand[0], cand[1]) > (best_candidate[0], best_candidate[1]):
            best_candidate = cand

    return best_candidate


def detect_layout(lines, keys):
    lengths = get_test_lengths(keys)
    best_two = _detect_layout_two_blocks(lines, keys, lengths)
    best_four = _detect_layout_four_blocks(lines, keys, lengths)

    if best_two is None and best_four is None:
        raise ValueError("TXT içinde cevap blokları bulunamadı.")
    if best_two is None:
        return best_four[2]
    if best_four is None:
        return best_two[2]
    return best_four[2] if (best_four[0], best_four[1]) > (best_two[0], best_two[1]) else best_two[2]


def parse_student_line(line, layout, lengths):
    prefix = line[:55]
    ans = extract_answers(line, layout, lengths)

    tc_match = re.search(r"\b\d{11}\b", prefix)
    tc_no = tc_match.group(0) if tc_match else ""

    if tc_no:
        name_part = prefix[:tc_match.start()]
        class_part = prefix[tc_match.end():]
    else:
        name_part = prefix
        class_part = ""

    # Remove symbols used in first columns.
    ad_soyad = re.sub(r"[*0-9]", " ", name_part)
    ad_soyad = re.sub(r"\s+", " ", ad_soyad).strip()
    sinif = class_part.strip()

    return tc_no, ad_soyad, sinif, ans


def score_with_booklet(ans_map, keyset):
    out = {}
    total = 0.0
    for test_name, _ in TESTS:
        d, y, b, n = calculate_score(ans_map[test_name], keyset[test_name])
        out[f"{test_name} Doğru"] = d
        out[f"{test_name} Yanlış"] = y
        out[f"{test_name} Boş"] = b
        out[f"{test_name} Net"] = n
        total += n
    out["Toplam Net"] = total
    return out


def d_y_n(ans, key):
    d, y, b, n = calculate_score(ans, key)
    return d, y, n


def compute_tyt_score(row):
    score = TYT_BASE_SCORE
    score += row["Türkçe Net"] * TYT_COEFFICIENTS["Türkçe"]
    score += row["Sosyal Net"] * TYT_COEFFICIENTS["Sosyal"]
    score += row["Matematik Net"] * TYT_COEFFICIENTS["Matematik"]
    score += row["Fen Net"] * TYT_COEFFICIENTS["Fen"]
    return round(score, 3)


def compute_weighted_score(row, base, coefs):
    score = base
    score += row["Türkçe Net"] * coefs["Türkçe"]
    score += row["Sosyal Net"] * coefs["Sosyal"]
    score += row["Matematik Net"] * coefs["Matematik"]
    score += row["Fen Net"] * coefs["Fen"]
    return round(score, 3)


def compute_soz_score(row):
    return compute_weighted_score(row, SOZ_BASE_SCORE, SOZ_COEFFICIENTS)


def compute_say_score(row):
    return compute_weighted_score(row, SAY_BASE_SCORE, SAY_COEFFICIENTS)


def compute_ea_score(row):
    return compute_weighted_score(row, EA_BASE_SCORE, EA_COEFFICIENTS)


def compute_ayt_score(row):
    return compute_weighted_score(row, AYT_BASE_SCORE, AYT_COEFFICIENTS)


def normalize_external_test_name(name):
    s = unicodedata.normalize("NFKD", str(name or ""))
    s = "".join(ch for ch in s if not unicodedata.combining(ch)).upper()
    if "EDEBIYAT-SOSYAL-1" in s or "EDEBIYAT SOSYAL-1" in s or "EDEBIYAT SOSYAL 1" in s:
        return "Türkçe"
    if "SOSYAL-2" in s or "SOSYAL 2" in s:
        return "Sosyal"
    if "SOSYAL" in s:
        return "Sosyal"
    if "MATEMATIK" in s:
        return "Matematik"
    if "FEN" in s:
        return "Fen"
    return None


def load_kazanim_table(path):
    def canonical(s):
        n = unicodedata.normalize("NFKD", str(s or ""))
        n = "".join(ch for ch in n if not unicodedata.combining(ch))
        n = n.lower().strip()
        n = n.replace("ı", "i")
        n = re.sub(r"\s+", " ", n)
        return n

    col_aliases = {
        "kitapcik": "Kitapçık",
        "test": "Test",
        "ders": "Ders",
        "a soru": "A Soru",
        "b soru": "B Soru",
        "cevap": "Cevap",
        "kazanim kodu": "Kazanım Kodu",
        "kazanım kodu": "Kazanım Kodu",
    }
    required = {"Kitapçık", "Test", "Ders", "A Soru", "B Soru", "Cevap", "Kazanım Kodu"}

    def standardize_columns(df):
        renamed = {}
        for c in df.columns:
            k = canonical(c)
            if k in col_aliases:
                renamed[c] = col_aliases[k]
        if renamed:
            df = df.rename(columns=renamed)
        df.columns = [str(c).strip() for c in df.columns]
        return df

    def has_required(df):
        cols = set(df.columns)
        return required.issubset(cols)

    def read_candidates(p):
        is_csv = p.endswith(".csv")
        if is_csv:
            readers = [
                lambda: pd.read_csv(path, sep=";", encoding="utf-8-sig", header=2),
                lambda: pd.read_csv(path, sep=";", encoding="utf-8-sig", header=0),
            ]
        else:
            readers = [
                lambda: pd.read_excel(path, header=2),
                lambda: pd.read_excel(path, header=0),
            ]
        for rd in readers:
            try:
                df = standardize_columns(rd())
                if has_required(df):
                    return df
            except Exception:
                pass
        return None

    p = str(path).lower()
    table = read_candidates(p)
    if table is None:
        # Last resort: detect header row from raw sheet by scanning first 20 rows.
        raw = pd.read_csv(path, sep=";", encoding="utf-8-sig", header=None) if p.endswith(".csv") else pd.read_excel(path, header=None)
        header_idx = None
        for i in range(min(20, len(raw))):
            vals = [canonical(v) for v in raw.iloc[i].tolist()]
            if {"kitapcik", "test", "ders", "a soru", "b soru", "cevap", "kazanim kodu"}.issubset(set(vals)):
                header_idx = i
                break
        if header_idx is None:
            raise ValueError("Kazanım tablosu başlık satırı okunamadı.")
        cols = [str(x).strip() for x in raw.iloc[header_idx].tolist()]
        table = raw.iloc[header_idx + 1 :].copy()
        table.columns = cols
        table = standardize_columns(table)
        if not has_required(table):
            missing = required.difference(set(table.columns))
            raise ValueError(f"Kazanım tablosunda eksik kolonlar var: {', '.join(sorted(missing))}")

    for col in ["A Soru", "B Soru"]:
        table[col] = pd.to_numeric(table[col], errors="coerce")

    kazanım_cols = [c for c in table.columns if c.startswith("Kazanım-")]
    for c in kazanım_cols:
        table[c] = table[c].fillna("").astype(str).str.strip()

    def build_kazanim_text(row):
        parts = [row.get("Kazanım Kodu", "")]
        parts.extend(row.get(c, "") for c in kazanım_cols)
        parts = [str(x).strip() for x in parts if str(x).strip() and str(x).strip().lower() != "nan"]
        return " | ".join(parts)

    table["__test_key"] = table["Test"].map(normalize_external_test_name)
    table["__kazanım_text"] = table.apply(build_kazanim_text, axis=1)
    table = table[table["__test_key"].notna()].copy()
    return table


def build_kazanim_analysis(df, keys, kazanim_table):
    rows = []
    for _, stu in df.iterrows():
        booklet = str(stu.get("Tahmini Kitapçık", "")).strip().upper()
        if booklet not in {"A", "B"}:
            continue

        q_col = "A Soru" if booklet == "A" else "B Soru"
        for _, kz in kazanim_table.iterrows():
            q = kz.get(q_col)
            if pd.isna(q):
                continue
            q = int(q)
            test_name = kz["__test_key"]
            ans_col = f"Ans {test_name}"
            if ans_col not in stu:
                continue
            ans = str(stu.get(ans_col, ""))
            key = str(keys.get(booklet, {}).get(test_name, ""))
            if q < 1 or q > len(ans) or q > len(key):
                continue
            if key[q - 1] == " ":
                continue

            student_mark = ans[q - 1]
            key_mark = key[q - 1]
            if student_mark == " ":
                durum = "Boş"
            elif student_mark == key_mark:
                durum = "Doğru"
            else:
                durum = "Yanlış"

            rows.append(
                {
                    "Ad Soyad": stu.get("Ad Soyad", ""),
                    "Sınıf": stu.get("Sınıf", ""),
                    "Kitapçık": booklet,
                    "Test": kz.get("Test", ""),
                    "Ders": kz.get("Ders", ""),
                    "Soru No": q,
                    "Doğru Cevap": key_mark,
                    "Öğrenci Cevap": student_mark,
                    "Durum": durum,
                    "Kazanım Kodu": kz.get("Kazanım Kodu", ""),
                    "Kazanım": kz.get("__kazanım_text", ""),
                }
            )

    detail = pd.DataFrame(rows)
    if detail.empty:
        return detail, detail, detail

    per_student = (
        detail.groupby(["Ad Soyad", "Sınıf", "Test", "Ders", "Kazanım Kodu", "Kazanım", "Durum"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for c in ["Doğru", "Yanlış", "Boş"]:
        if c not in per_student.columns:
            per_student[c] = 0
    per_student["Soru Sayısı"] = per_student["Doğru"] + per_student["Yanlış"] + per_student["Boş"]
    per_student["Başarı %"] = (
        (per_student["Doğru"] / per_student["Soru Sayısı"]).replace([float("inf")], 0).fillna(0) * 100.0
    ).round(1)
    weak = per_student[(per_student["Yanlış"] + per_student["Boş"] > 0)].sort_values(
        ["Başarı %", "Yanlış", "Boş"], ascending=[True, False, False]
    )

    class_summary = (
        detail.groupby(["Test", "Ders", "Kazanım Kodu", "Kazanım", "Durum"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for c in ["Doğru", "Yanlış", "Boş"]:
        if c not in class_summary.columns:
            class_summary[c] = 0
    class_summary["Soru Sayısı"] = (
        class_summary["Doğru"] + class_summary["Yanlış"] + class_summary["Boş"]
    )
    class_summary["Başarı %"] = (
        (class_summary["Doğru"] / class_summary["Soru Sayısı"]).replace([float("inf")], 0).fillna(0) * 100.0
    ).round(1)
    class_summary = class_summary.sort_values(["Başarı %", "Yanlış", "Boş"], ascending=[True, False, False])

    return detail, weak, class_summary


def run_pipeline(txt_path, pdf_path, output_xlsx, kazanim_path=None):
    print("Cevap anahtarı okunuyor...")
    if str(pdf_path).lower().endswith(".json"):
        keys = load_standard_key_json(pdf_path)
    else:
        keys = normalize_keys(parse_pdf_keys(pdf_path))

    per_booklet_quality = {
        booklet: sum(
            sum(1 for ch in keys.get(booklet, {}).get(test_name, "") if ch in "ABCDE")
            for test_name, _ in TESTS
        )
        for booklet in ["A", "B"]
    }
    if max(per_booklet_quality.values()) < 80:
        raise ValueError(
            "PDF cevap anahtarı yeterli doğrulukta okunamadı. "
            "Muhtemelen farklı başlık/yerleşim var; lütfen kontrol edin."
        )
    for booklet in ["A", "B"]:
        lens = {k: len(v) for k, v in keys[booklet].items()}
        print(f"{booklet} anahtarı uzunlukları: {lens}")

    print("Öğrenci yanıtları okunuyor...")
    results = []
    with open(txt_path, "r", encoding="windows-1254", errors="replace") as f:
        lines = [line.rstrip("\n") for line in f if line.strip()]

    lengths = get_test_lengths(keys)
    layout = detect_layout(lines, keys)
    if layout["mode"] == "four_blocks":
        print(f"Tespit edilen layout (4 blok): starts={layout['starts']}")
    else:
        print(
            "Tespit edilen layout (2 blok): "
            f"s1={layout['s1']}, s2={layout['s2']}, o1={layout['o1']}, o2={layout['o2']}"
        )

    for line in lines:
        if layout["mode"] == "four_blocks":
            t_start, s_start, m_start, f_start = layout["starts"]
            min_len = max(
                t_start + lengths["Türkçe"],
                s_start + lengths["Sosyal"],
                m_start + lengths["Matematik"],
                f_start + lengths["Fen"],
            )
        else:
            g1 = lengths["Türkçe"] + lengths["Sosyal"]
            g2 = lengths["Matematik"] + lengths["Fen"]
            min_len = max(layout["s1"] + g1, layout["s2"] + g2)

        if len(line) < min_len:
            continue

        tc_no, ad_soyad, sinif, answers = parse_student_line(line, layout, lengths)
        score_a = score_with_booklet(answers, keys["A"])
        score_b = score_with_booklet(answers, keys["B"])

        best_booklet = "A" if score_a["Toplam Net"] >= score_b["Toplam Net"] else "B"
        best_score = score_a if best_booklet == "A" else score_b

        results.append(
            {
                "TC No": tc_no,
                "Ad Soyad": ad_soyad,
                "Sınıf": sinif,
                "Tahmini Kitapçık": best_booklet,
                "A Toplam Net": score_a["Toplam Net"],
                "B Toplam Net": score_b["Toplam Net"],
                "Ans Türkçe": answers["Türkçe"],
                "Ans Sosyal": answers["Sosyal"],
                "Ans Matematik": answers["Matematik"],
                "Ans Fen": answers["Fen"],
                **best_score,
            }
        )

    df = pd.DataFrame(results)
    if df.empty:
        print("Veri okunamadı.")
        return

    df = df.sort_values("Toplam Net", ascending=False).reset_index(drop=True)
    df.insert(0, "Sıra", range(1, len(df) + 1))
    df["TYT Puan"] = df.apply(compute_tyt_score, axis=1)
    df["AYT Puan"] = df.apply(compute_ayt_score, axis=1)
    df["SÖZ Puan"] = df.apply(compute_soz_score, axis=1)
    df["SAY Puan"] = df.apply(compute_say_score, axis=1)
    df["EA Puan"] = df.apply(compute_ea_score, axis=1)
    df["TYT Sıra"] = (
        df["TYT Puan"]
        .rank(method="min", ascending=False)
        .astype(int)
    )
    df["SÖZ Sıra"] = (
        df["SÖZ Puan"]
        .rank(method="min", ascending=False)
        .astype(int)
    )
    df["SAY Sıra"] = (
        df["SAY Puan"]
        .rank(method="min", ascending=False)
        .astype(int)
    )
    df["EA Sıra"] = (
        df["EA Puan"]
        .rank(method="min", ascending=False)
        .astype(int)
    )
    df["AYT Sıra"] = (
        df["AYT Puan"]
        .rank(method="min", ascending=False)
        .astype(int)
    )

    stats_rows = []
    for sinif, group in df.groupby("Sınıf"):
        stats_rows.append(
            {
                "Sınıf/Şube": sinif if sinif else "Belirtilmemiş",
                "Öğrenci Sayısı": len(group),
                "Türkçe Ort. Net": group["Türkçe Net"].mean(),
                "Sosyal Ort. Net": group["Sosyal Net"].mean(),
                "Matematik Ort. Net": group["Matematik Net"].mean(),
                "Fen Ort. Net": group["Fen Net"].mean(),
                "Toplam Ort. Net": group["Toplam Net"].mean(),
            }
        )
    stats_rows.append(
        {
            "Sınıf/Şube": "GENEL ORTALAMA",
            "Öğrenci Sayısı": len(df),
            "Türkçe Ort. Net": df["Türkçe Net"].mean(),
            "Sosyal Ort. Net": df["Sosyal Net"].mean(),
            "Matematik Ort. Net": df["Matematik Net"].mean(),
            "Fen Ort. Net": df["Fen Net"].mean(),
            "Toplam Ort. Net": df["Toplam Net"].mean(),
        }
    )
    df_stats = pd.DataFrame(stats_rows)

    exam_profile = detect_exam_profile(lengths)
    print(f"Sınav profili: {exam_profile}")

    report_rows = []
    if exam_profile == "school_30x4":
        for _, r in df.iterrows():
            toplam_d = r["Türkçe Doğru"] + r["Sosyal Doğru"] + r["Matematik Doğru"] + r["Fen Doğru"]
            toplam_y = r["Türkçe Yanlış"] + r["Sosyal Yanlış"] + r["Matematik Yanlış"] + r["Fen Yanlış"]
            report_rows.append(
                {
                    "Sıra": int(r["Sıra"]),
                    "Numara": 0,
                    "İsim": r["Ad Soyad"],
                    "Sınıf": r["Sınıf"] if pd.notna(r["Sınıf"]) else "",
                    "TYT Türkçe D": r["Türkçe Doğru"],
                    "TYT Türkçe Y": r["Türkçe Yanlış"],
                    "TYT Türkçe N": r["Türkçe Net"],
                    "TYT Sosyal D": r["Sosyal Doğru"],
                    "TYT Sosyal Y": r["Sosyal Yanlış"],
                    "TYT Sosyal N": r["Sosyal Net"],
                    "TYT Matematik D": r["Matematik Doğru"],
                    "TYT Matematik Y": r["Matematik Yanlış"],
                    "TYT Matematik N": r["Matematik Net"],
                    "TYT Fen D": r["Fen Doğru"],
                    "TYT Fen Y": r["Fen Yanlış"],
                    "TYT Fen N": r["Fen Net"],
                    "Toplam D": toplam_d,
                    "Toplam Y": toplam_y,
                    "NET": r["Toplam Net"],
                    "TYT Puan": r["TYT Puan"],
                    "TYT Sıra": int(r["TYT Sıra"]),
                    "TYT Okul": int(r["TYT Sıra"]),
                    "TYT Genel": int(r["TYT Sıra"]),
                    "SÖZ Puan": r["SÖZ Puan"],
                    "SÖZ Sıra": int(r["SÖZ Sıra"]),
                    "SÖZ Okul": int(r["SÖZ Sıra"]),
                    "SÖZ Genel": int(r["SÖZ Sıra"]),
                    "SAY Puan": r["SAY Puan"],
                    "SAY Sıra": int(r["SAY Sıra"]),
                    "SAY Okul": int(r["SAY Sıra"]),
                    "SAY Genel": int(r["SAY Sıra"]),
                    "EA Puan": r["EA Puan"],
                    "EA Sıra": int(r["EA Sıra"]),
                    "EA Okul": int(r["EA Sıra"]),
                    "EA Genel": int(r["EA Sıra"]),
                }
            )
    elif exam_profile == "tyt":
        for _, r in df.iterrows():
            bk = r["Tahmini Kitapçık"]
            key = keys[bk]
            sos = r["Ans Sosyal"]
            fen = r["Ans Fen"]
            tar_d, tar_y, tar_n = d_y_n(sos[0:5], key["Sosyal"][0:5])
            cog_d, cog_y, cog_n = d_y_n(sos[5:10], key["Sosyal"][5:10])
            fel_d, fel_y, fel_n = d_y_n(sos[10:15], key["Sosyal"][10:15])
            din_d, din_y, din_n = d_y_n(sos[15:20], key["Sosyal"][15:20])
            fiz_d, fiz_y, fiz_n = d_y_n(fen[0:7], key["Fen"][0:7])
            kim_d, kim_y, kim_n = d_y_n(fen[7:14], key["Fen"][7:14])
            bio_d, bio_y, bio_n = d_y_n(fen[14:20], key["Fen"][14:20])
            toplam_d = r["Türkçe Doğru"] + r["Sosyal Doğru"] + r["Matematik Doğru"] + r["Fen Doğru"]
            toplam_y = r["Türkçe Yanlış"] + r["Sosyal Yanlış"] + r["Matematik Yanlış"] + r["Fen Yanlış"]
            report_rows.append(
                {
                    "Sıra": int(r["Sıra"]),
                    "Numara": 0,
                    "İsim": r["Ad Soyad"],
                    "Sınıf": r["Sınıf"] if pd.notna(r["Sınıf"]) else "",
                    "Türkçe D": r["Türkçe Doğru"], "Türkçe Y": r["Türkçe Yanlış"], "Türkçe N": r["Türkçe Net"],
                    "Tarih D": tar_d, "Tarih Y": tar_y, "Tarih N": tar_n,
                    "Coğrafya D": cog_d, "Coğrafya Y": cog_y, "Coğrafya N": cog_n,
                    "Felsefe D": fel_d, "Felsefe Y": fel_y, "Felsefe N": fel_n,
                    "Din D": din_d, "Din Y": din_y, "Din N": din_n,
                    "Matematik D": r["Matematik Doğru"], "Matematik Y": r["Matematik Yanlış"], "Matematik N": r["Matematik Net"],
                    "Fizik D": fiz_d, "Fizik Y": fiz_y, "Fizik N": fiz_n,
                    "Kimya D": kim_d, "Kimya Y": kim_y, "Kimya N": kim_n,
                    "Biyoloji D": bio_d, "Biyoloji Y": bio_y, "Biyoloji N": bio_n,
                    "Toplam D": toplam_d, "Toplam Y": toplam_y, "NET": r["Toplam Net"],
                    "TYT Puan": r["TYT Puan"],
                    "TYT Sıra": int(r["TYT Sıra"]),
                    "TYT Okul": int(r["TYT Sıra"]),
                    "TYT Genel": int(r["TYT Sıra"]),
                }
            )
    else:
        for _, r in df.iterrows():
            toplam_d = r["Türkçe Doğru"] + r["Sosyal Doğru"] + r["Matematik Doğru"] + r["Fen Doğru"]
            toplam_y = r["Türkçe Yanlış"] + r["Sosyal Yanlış"] + r["Matematik Yanlış"] + r["Fen Yanlış"]
            report_rows.append(
                {
                    "Sıra": int(r["Sıra"]),
                    "Numara": 0,
                    "İsim": r["Ad Soyad"],
                    "Sınıf": r["Sınıf"] if pd.notna(r["Sınıf"]) else "",
                    "AYT TDE-S1 D": r["Türkçe Doğru"], "AYT TDE-S1 Y": r["Türkçe Yanlış"], "AYT TDE-S1 N": r["Türkçe Net"],
                    "AYT Sosyal-2 D": r["Sosyal Doğru"], "AYT Sosyal-2 Y": r["Sosyal Yanlış"], "AYT Sosyal-2 N": r["Sosyal Net"],
                    "AYT Matematik D": r["Matematik Doğru"], "AYT Matematik Y": r["Matematik Yanlış"], "AYT Matematik N": r["Matematik Net"],
                    "AYT Fen D": r["Fen Doğru"], "AYT Fen Y": r["Fen Yanlış"], "AYT Fen N": r["Fen Net"],
                    "Toplam D": toplam_d, "Toplam Y": toplam_y, "NET": r["Toplam Net"],
                    "AYT Puan": r["AYT Puan"],
                    "AYT Sıra": int(r["AYT Sıra"]),
                    "AYT Okul": int(r["AYT Sıra"]),
                    "AYT Genel": int(r["AYT Sıra"]),
                }
            )

    df_report = pd.DataFrame(report_rows)
    avg_row = {"Sıra": "", "Numara": "", "İsim": "Genel Ortalama", "Sınıf": ""}
    numeric_cols = [c for c in df_report.columns if c not in {"Sıra", "Numara", "İsim", "Sınıf"}]
    for col in numeric_cols:
        series = pd.to_numeric(df_report[col], errors="coerce")
        avg_row[col] = round(series.mean(), 3) if series.notna().any() else ""
    for col in df_report.columns:
        avg_row.setdefault(col, "")
    df_report = pd.concat([pd.DataFrame([avg_row]), df_report], ignore_index=True)

    df_net = df_report.copy()
    student_mask = df_net["Sıra"] != ""
    students = df_net[student_mask].copy().sort_values("NET", ascending=False).reset_index(drop=True)
    students["Sıra"] = range(1, len(students) + 1)
    df_net = pd.concat([df_net[~student_mask], students], ignore_index=True)

    kazanim_detail = pd.DataFrame()
    kazanim_weak = pd.DataFrame()
    kazanim_summary = pd.DataFrame()
    if kazanim_path:
        kz_table = load_kazanim_table(kazanim_path)
        kazanim_detail, kazanim_weak, kazanim_summary = build_kazanim_analysis(df, keys, kz_table)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_report.to_excel(writer, sheet_name="TYT", index=False)
        df_net.to_excel(writer, sheet_name="NET SIRALI", index=False)
        df.to_excel(writer, sheet_name="Sonuçlar", index=False)
        df_stats.to_excel(writer, sheet_name="İstatistikler", index=False)
        if not kazanim_weak.empty:
            kazanim_weak.to_excel(writer, sheet_name="Kazanım Zayıf", index=False)
        if not kazanim_summary.empty:
            kazanim_summary.to_excel(writer, sheet_name="Kazanım Özet", index=False)
        if not kazanim_detail.empty:
            kazanim_detail.to_excel(writer, sheet_name="Kazanım Detay", index=False)

        def style_school(ws):
            ws.insert_rows(1, amount=5)
            max_col = ws.max_column
            ws["A1"] = "OKUL TEST BAZLI NET LİSTESİ"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws["A2"] = "İL"; ws["D2"] = "İLÇE"; ws["G2"] = "OKUL"; ws["P2"] = "SINAV ADI"; ws["AF2"] = "KATILIMLAR"
            ws.merge_cells("A2:C2"); ws.merge_cells("D2:F2"); ws.merge_cells("G2:O2"); ws.merge_cells("P2:AE2"); ws.merge_cells("AF2:AI2")
            ws["AF3"] = "Okul"; ws["AG3"] = "İlçe"; ws["AH3"] = "İl"; ws["AI3"] = "Genel"; ws["AF4"] = len(df)
            ws.merge_cells("A3:C4"); ws.merge_cells("D3:F4"); ws.merge_cells("G3:O4"); ws.merge_cells("P3:AE4")
            ws["E5"] = "TYT Türkçe"; ws["H5"] = "TYT Sosyal"; ws["K5"] = "TYT Matematik"; ws["N5"] = "TYT Fen"; ws["Q5"] = "Toplam"
            ws["T5"] = "TYT"; ws["X5"] = "SÖZ"; ws["AB5"] = "SAY"; ws["AF5"] = "EA"
            ws.merge_cells("E5:G5"); ws.merge_cells("H5:J5"); ws.merge_cells("K5:M5"); ws.merge_cells("N5:P5"); ws.merge_cells("Q5:S5")
            ws.merge_cells("T5:W5"); ws.merge_cells("X5:AA5"); ws.merge_cells("AB5:AE5"); ws.merge_cells("AF5:AI5")
            for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=max_col):
                for c in row:
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=max_col):
                for c in row:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        def style_tyt(ws):
            ws.insert_rows(1, amount=2)
            ws["E1"] = "TYT-TÜRKÇE"; ws["H1"] = "TYT-SOSYAL BİLİMLER"; ws["T1"] = "TYT-MATEMATİK"; ws["W1"] = "TYT-FEN BİLİMLERİ"
            ws["AF1"] = "TOPLAM"; ws["AI1"] = "TYT"
            ws.merge_cells("E1:G1"); ws.merge_cells("H1:S1"); ws.merge_cells("T1:V1"); ws.merge_cells("W1:AE1"); ws.merge_cells("AF1:AH1"); ws.merge_cells("AI1:AL1")
            for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=ws.max_column):
                for c in row:
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for c in row:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        def style_ayt(ws):
            ws.insert_rows(1, amount=2)
            ws["E1"] = "AYT TDE-S1"; ws["H1"] = "AYT SOSYAL-2"; ws["K1"] = "AYT MATEMATİK"; ws["N1"] = "AYT FEN"; ws["Q1"] = "TOPLAM"; ws["T1"] = "AYT"
            ws.merge_cells("E1:G1"); ws.merge_cells("H1:J1"); ws.merge_cells("K1:M1"); ws.merge_cells("N1:P1"); ws.merge_cells("Q1:S1"); ws.merge_cells("T1:W1")
            for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=ws.max_column):
                for c in row:
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for c in row:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        if exam_profile == "school_30x4":
            style_school(writer.book["TYT"])
            style_school(writer.book["NET SIRALI"])
        elif exam_profile == "tyt":
            style_tyt(writer.book["TYT"])
            style_tyt(writer.book["NET SIRALI"])
        else:
            style_ayt(writer.book["TYT"])
            style_ayt(writer.book["NET SIRALI"])

    print(f"Tamamlandı: {output_xlsx}")
    print("Tahmini kitapçık dağılımı:")
    print(df["Tahmini Kitapçık"].value_counts().to_string())
    print(f"En yüksek net: {df['Toplam Net'].max():.2f}")

    # Save canonical answer-key next to output for reproducible re-runs.
    key_json_path = str(output_xlsx).rsplit(".", 1)[0] + "_answer_key.standard.json"
    save_standard_key_json(keys, key_json_path)
    print(f"Standart cevap anahtarı kaydedildi: {key_json_path}")
    if not kazanim_weak.empty:
        print(f"Kazanım analizi üretildi: {len(kazanim_weak)} satır zayıf kazanım.")
        df.attrs["kazanim_weak_preview"] = kazanim_weak.head(20).to_dict("records")
    return df


def main():
    run_pipeline(TXT_PATH, PDF_PATH, OUTPUT_XLSX)


if __name__ == "__main__":
    main()
